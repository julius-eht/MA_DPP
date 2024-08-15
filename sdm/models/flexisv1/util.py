from openpyxl import load_workbook
import pandas as pd
from pydantic import BaseModel
from aas2openapi.util.convert_util import convert_camel_case_to_underscrore_str

class FlexisDataFrames(BaseModel):
    job_shop_scheduling_job: pd.DataFrame
    job_shop_scheduling_job_task: pd.DataFrame
    resource_task: pd.DataFrame
    resource_variant: pd.DataFrame
    resource_availibility: pd.DataFrame
    job_shop_scheduling_scheduled_task: pd.DataFrame
    orders: pd.DataFrame
    process_times: pd.DataFrame
    setup_times: pd.DataFrame
    tasks: pd.DataFrame
    transport_times: pd.DataFrame
    work_plan: pd.DataFrame
    work_plan_task: pd.DataFrame
    connections: pd.DataFrame
    movement_mode: pd.DataFrame
    resources: pd.DataFrame
    resource_types: pd.DataFrame
    resource_positions: pd.DataFrame
    scenarios: pd.DataFrame

    class Config:
        arbitrary_types_allowed = True

    

def read_excel(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    data = {}
    sheet_names = list(FlexisDataFrames.__annotations__.keys())
    for sheet, sheet_name in zip(sheets, sheet_names):
        data[sheet_name] = pd.DataFrame(
            wb[sheet].values,
            columns=[cell.value for cell in wb[sheet][1]],
        )
        data[sheet_name] = data[sheet_name].dropna(how="all")
        data[sheet_name] = data[sheet_name].iloc[1:]
        new_columns = [column for column in data[sheet_name].columns if column]
        data[sheet_name] = data[sheet_name][new_columns]
    return data